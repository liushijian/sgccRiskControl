# -*- coding: utf-8 -*-
import os
from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import PatternFill,Color, Style

class DataProcess():
    def __init__(self,root):
        self.root=root
    def split_table(self):
        for fileName in os.listdir(self.root):
            if fileName.endswith(".xlsx"):
                f=os.path.join(self.root,fileName)
                wb=load_workbook(f)
                for sheetName in wb.get_sheet_names():
                    if sheetName!="Knight":
                        df=pd.read_excel(f,sheetname=sheetName)
                        companies=df["稽核对象"].drop_duplicates().iteritems()
                        for company in companies:
                            dirName=os.path.join(self.root,company[1])
                            if not os.path.exists(dirName):
                                os.mkdir(dirName)
                            filePath=os.path.join(dirName,fileName)
                            df[df["稽核对象"]==company[1]].to_excel(filePath)
    def drop_absEqual(self):
        for f in os.listdir(self.root):
            f = os.path.join(self.root, f)
            wb = load_workbook(f)
            for sheetName in wb.get_sheet_names():
                if sheetName != "Knight":
                    ws = wb[sheetName]
                    df=pd.read_excel(f,sheetname=sheetName)
                    for i in range(2,ws.max_row+1):
                        money=float(ws.cell(row=i,column=9).value)
                        df["value"]=df["期末余额"]+money
                        df=df[df["value"].abs()<0.1]
                        if len(df)>0:
                            print (f,i,df)
                            ws.row_dimensions[i].style = \
                                Style(fill=PatternFill(patternType='solid', fgColor=Color('FFFF0000')))
            wb.save(f)



zhiGongsi = "E:\\大数据北京\\code\\拆分\\省公司、直属单位"
yiyi="E:\\大数据北京\\code\\一正一负"
d=DataProcess(zhiGongsi)
d.split_table()
